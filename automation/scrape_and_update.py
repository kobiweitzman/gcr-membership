#!/usr/bin/env python3
"""
Automated BBYO membership data scraper and uploader.

Uses Playwright to:
  1. Log in to bbyo.my.site.com
  2. Navigate to My Chapter
  3. Export membership data as Excel
  4. Convert to JSON (stripping PII)
  5. Upload to Supabase

Environment variables required:
  BBYO_EMAIL              - BBYO login email
  BBYO_PASSWORD           - BBYO login password
  SUPABASE_URL            - Supabase project URL
  SUPABASE_SERVICE_ROLE_KEY - Supabase service role key
"""

import json
import os
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Installing playwright...")
    os.system(f"{sys.executable} -m pip install playwright -q")
    os.system(f"{sys.executable} -m playwright install chromium")
    from playwright.sync_api import sync_playwright


# ---------- Configuration ----------

BBYO_EMAIL = os.environ.get("BBYO_EMAIL", "")
BBYO_PASSWORD = os.environ.get("BBYO_PASSWORD", "")
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

LOGIN_URL = "https://bbyo.my.site.com/s/login"
MY_CHAPTER_URL = "https://bbyo.my.site.com/s/my-chapter"

SENSITIVE_FIELDS = [
    "phone", "email",
    "parent1Name", "parent1Email", "parent1Cell",
    "parent2Name", "parent2Email", "parent2Cell",
]


# ---------- Scraping ----------

def scrape_membership_export():
    """Use Playwright to log in and export the membership Excel file."""
    print("Launching browser...")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        # --- Login ---
        print("Navigating to login page...")
        # Use domcontentloaded — Salesforce never reaches networkidle
        page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=60000)

        # Wait for the login form to actually render
        print("Waiting for login form...")
        page.wait_for_selector("input[type='password']", timeout=30000)
        page.wait_for_timeout(1000)

        print("Filling login form...")
        # Find and fill email field
        email_input = page.locator("input[type='email'], input[name='username'], input[placeholder*='mail']").first
        email_input.fill(BBYO_EMAIL)

        # Find and fill password field
        password_input = page.locator("input[type='password']").first
        password_input.fill(BBYO_PASSWORD)

        # Click Log In button
        login_btn = page.locator("button:has-text('Log In'), button:has-text('Login'), input[type='submit']").first
        login_btn.click()

        print("Logging in... waiting for redirect...")
        page.wait_for_timeout(5000)

        # Wait for login to complete - look for the navigation bar
        try:
            page.wait_for_selector("text=My Chapter", timeout=30000)
            print("Login successful!")
        except Exception:
            # Try waiting a bit more
            page.wait_for_timeout(5000)
            if "login" in page.url.lower():
                print("ERROR: Login appears to have failed. Check credentials.")
                print(f"Current URL: {page.url}")
                page.screenshot(path="/tmp/debug_login_fail.png")
                browser.close()
                sys.exit(1)
            print("Login appears successful (redirected away from login page).")

        # --- Navigate to My Chapter ---
        print("Navigating to My Chapter...")
        # Use domcontentloaded — Salesforce never reaches networkidle
        page.goto(MY_CHAPTER_URL, wait_until="domcontentloaded", timeout=60000)
        print("Page DOM loaded, waiting for content to render...")

        # Wait for the member table to appear (more reliable than fixed timeout)
        try:
            page.wait_for_selector("text=All Members in My Chapter", timeout=45000)
            print("My Chapter page loaded.")
        except Exception:
            print("Warning: 'All Members' text not found, trying alternate selectors...")
            try:
                page.wait_for_selector("button:has-text('Export')", timeout=30000)
                print("Found Export button — page is loaded.")
            except Exception:
                print("Warning: Could not confirm page load. Taking debug screenshot...")
                page.screenshot(path="/tmp/debug_my_chapter.png")
                print(f"Current URL: {page.url}")
                page.wait_for_timeout(5000)

        # Wait for actual data rows to load (not just the header)
        print("Waiting for member data to load in table...")
        try:
            # Look for table rows or data cells that indicate members have loaded
            page.wait_for_selector(
                "table tbody tr, "                         # standard table rows
                "[class*='datatable'] [role='row'], "      # SLDS datatable rows
                "[class*='slds-table'] tbody tr, "         # SLDS table rows
                "[data-row-key-value]",                    # Lightning datatable rows
                timeout=30000
            )
            page.wait_for_timeout(2000)  # Extra buffer for all rows to render
            print("Table data loaded.")
        except Exception:
            print("Warning: Could not detect table rows. Waiting extra time...")
            page.wait_for_timeout(10000)
            page.screenshot(path="/tmp/debug_no_table_rows.png")

        # --- Click Export button (the one on the page, not in dialog) ---
        print("Clicking Export button...")
        export_btn = page.locator("button.avonni-button:has-text('Export'), button:has-text('Export')").first
        try:
            export_btn.wait_for(timeout=15000)
        except Exception:
            print("Export button not found. Taking debug screenshot...")
            page.screenshot(path="/tmp/debug_no_export_btn.png")
            browser.close()
            sys.exit(1)

        export_btn.click()
        page.wait_for_timeout(2000)

        # --- Wait for Export dialog ---
        print("Waiting for export dialog...")
        try:
            page.wait_for_selector("text=Export View", timeout=15000)
            print("Export dialog opened.")
        except Exception:
            print("Warning: Export dialog may not have opened properly.")
            page.screenshot(path="/tmp/debug_export_dialog.png")
            page.wait_for_timeout(3000)

        # --- Select "Details Only" so ALL rows are exported ---
        # The dialog defaults to "Formatted Report", which only exports the
        # ~30 rows currently rendered in the report preview. "Details Only"
        # exports every row. Without this, we get 30 members instead of ~994.
        debug_dir = os.environ.get("DEBUG_ARTIFACT_DIR", "/tmp")

        try:
            page.screenshot(path=f"{debug_dir}/debug_before_details_click.png", full_page=True)
        except Exception:
            pass

        # Starting from the "Details Only" text (which Playwright finds by
        # piercing shadow DOM), walk UP via getRootNode().host to escape nested
        # shadow roots, then serialize from the outermost ancestor downwards
        # including all shadow subtrees.
        try:
            full_dialog_html = page.get_by_text("Details Only", exact=True).first.evaluate("""el => {
                // Walk up through shadow boundaries to find the outermost shadow host.
                let node = el;
                let hosts = [];
                while (node) {
                    const root = node.getRootNode();
                    if (root instanceof ShadowRoot) {
                        hosts.push(root.host);
                        node = root.host;
                    } else {
                        break;
                    }
                }
                // Take the highest ancestor (outermost shadow host) and go up
                // a few more levels in the light DOM to capture the whole modal.
                let anchor = hosts.length ? hosts[hosts.length - 1] : el;
                for (let i = 0; i < 6 && anchor.parentElement; i++) {
                    anchor = anchor.parentElement;
                }
                function serialize(el, depth) {
                    if (depth > 15) return '<T>';
                    if (!el || !el.tagName) return '';
                    let html = '<' + el.tagName.toLowerCase();
                    for (const attr of el.attributes || []) {
                        html += ' ' + attr.name + '="' + (attr.value||'').replace(/"/g,'&quot;').slice(0,120) + '"';
                    }
                    html += '>';
                    if (el.shadowRoot) {
                        html += '<!--SHADOW-->';
                        for (const c of el.shadowRoot.children) html += serialize(c, depth+1);
                        html += '<!--/SHADOW-->';
                    }
                    for (const c of el.childNodes || []) {
                        if (c.nodeType === 1) html += serialize(c, depth+1);
                        else if (c.nodeType === 3 && c.textContent.trim()) html += c.textContent.trim().slice(0,100);
                    }
                    html += '</' + el.tagName.toLowerCase() + '>';
                    return html;
                }
                return serialize(anchor, 0);
            }""")
            with open(f"{debug_dir}/debug_full_dialog.html", "w") as f:
                f.write(full_dialog_html)
            print(f"Saved full dialog with shadow subtrees ({len(full_dialog_html)} bytes)")
        except Exception as e:
            print(f"Full dialog dump failed: {e}")

        # Also capture all radio/checkbox/input/select names inside the dialog.
        try:
            inputs_summary = page.get_by_text("Details Only", exact=True).first.evaluate("""el => {
                let node = el;
                while (node) {
                    const root = node.getRootNode();
                    if (root instanceof ShadowRoot) node = root.host;
                    else break;
                }
                let anchor = node;
                for (let i=0;i<6 && anchor.parentElement;i++) anchor = anchor.parentElement;
                function* walk(root, depth) {
                    if (!root || depth > 15) return;
                    for (const c of (root.children||[])) {
                        if (c.tagName === 'INPUT' || c.tagName === 'SELECT' || c.tagName === 'BUTTON') {
                            yield {
                                tag: c.tagName,
                                type: c.type,
                                name: c.name,
                                value: (c.value||'').slice(0,60),
                                aria: c.getAttribute('aria-label'),
                                text: (c.textContent||'').trim().slice(0,60),
                                checked: c.checked,
                                disabled: c.disabled
                            };
                        }
                        yield* walk(c, depth+1);
                        if (c.shadowRoot) yield* walk(c.shadowRoot, depth+1);
                    }
                }
                return JSON.stringify(Array.from(walk(anchor, 0)), null, 2);
            }""")
            with open(f"{debug_dir}/debug_inputs.json", "w") as f:
                f.write(inputs_summary)
            print(f"Saved inputs summary ({len(inputs_summary)} bytes)")
        except Exception as e:
            print(f"Inputs dump failed: {e}")

        # Dump the HTML of the ancestor chain of the "Details Only" text using
        # Playwright's locator (which pierces Shadow DOM).
        try:
            details_text_loc = page.get_by_text("Details Only", exact=True).first
            if details_text_loc.count() > 0:
                ancestor_html = details_text_loc.evaluate(
                    "el => { let n = el; for (let i=0;i<12 && n;i++) {"
                    " if (n.querySelector && (n.querySelector('input[type=\"radio\"]')"
                    " || n.tagName==='LABEL')) return n.outerHTML.slice(0, 4000); n = n.parentElement; }"
                    " return el.outerHTML; }"
                )
                with open(f"{debug_dir}/debug_details_ancestor.html", "w") as f:
                    f.write(ancestor_html)
                print(f"Saved Details Only ancestor HTML ({len(ancestor_html)} bytes)")
            else:
                print("Could not find 'Details Only' text via get_by_text.")
        except Exception as e:
            print(f"Ancestor dump failed: {e}")

        print("Selecting 'Details Only' export option...")
        clicked_via = None

        # Strategy 1: radio role
        try:
            radio = page.get_by_role("radio", name="Details Only")
            if radio.count() > 0:
                radio.first.check(timeout=5000)
                clicked_via = "role=radio check()"
        except Exception as e:
            print(f"  role=radio failed: {e}")

        # Strategy 2: click the label ancestor of the text
        if not clicked_via:
            try:
                result = page.get_by_text("Details Only", exact=True).first.evaluate(
                    "el => { let n = el; for (let i=0;i<12 && n;i++) {"
                    " if (n.tagName==='LABEL') { n.click(); return 'LABEL'; }"
                    " const r = n.querySelector && n.querySelector('input[type=\"radio\"]');"
                    " if (r) { r.click(); return 'RADIO-INPUT'; } n = n.parentElement; } return null; }"
                )
                if result:
                    clicked_via = f"ancestor-click: {result}"
            except Exception as e:
                print(f"  ancestor-click failed: {e}")

        # Strategy 3: click the parent figure element
        if not clicked_via:
            try:
                page.get_by_text("Details Only", exact=True).first.locator(
                    "xpath=ancestor::*[contains(@class,'visual-picker') or contains(@class,'slds-visual-picker')][1]"
                ).click(timeout=5000)
                clicked_via = "xpath ancestor visual-picker"
            except Exception as e:
                print(f"  visual-picker ancestor failed: {e}")

        # Strategy 4: force-click the text itself
        if not clicked_via:
            try:
                page.get_by_text("Details Only", exact=True).first.click(force=True, timeout=5000)
                clicked_via = "force click text"
            except Exception as e:
                print(f"  force-click failed: {e}")

        if clicked_via:
            print(f"  Clicked 'Details Only' via: {clicked_via}")
        else:
            print("  WARNING: All 'Details Only' click strategies failed.")

        page.wait_for_timeout(1500)

        try:
            page.screenshot(path=f"{debug_dir}/debug_after_details_click.png", full_page=True)
        except Exception:
            pass

        # --- Enable "Export all data" toggle ---
        # Without this, the export is capped to the 30-row preview even when
        # "Details Only" is selected.
        print("Enabling 'Export all data' toggle...")
        try:
            toggle_label = page.get_by_text("Export all data", exact=True).first
            if toggle_label.count() > 0:
                toggle_label.click(timeout=5000)
                print("  Clicked 'Export all data' label.")
            else:
                print("  WARNING: 'Export all data' label not found.")
        except Exception as e:
            print(f"  WARNING: Export-all-data click failed: {e}")

        page.wait_for_timeout(500)

        try:
            page.screenshot(path=f"{debug_dir}/debug_after_export_all_click.png", full_page=True)
        except Exception:
            pass

        # --- Click the Export button inside the dialog (the brand/primary button) ---
        print("Clicking Export in dialog...")

        # Start waiting for download BEFORE clicking
        with page.expect_download(timeout=60000) as download_info:
            # Click the brand export button in the modal
            dialog_export_btn = page.locator("button.slds-button_brand:has-text('Export')").first
            dialog_export_btn.click()

        download = download_info.value
        print(f"Download started: {download.suggested_filename}")

        # Save to temp path
        download_path = f"/tmp/bbyo_export_{int(time.time())}.xlsx"
        download.save_as(download_path)
        print(f"File saved to: {download_path}")

        browser.close()
        return download_path


# ---------- Spreadsheet conversion ----------

def convert_spreadsheet(input_path):
    """Convert the Excel export to a list of member dictionaries."""
    print(f"\nProcessing: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    print(f"Found {ws.max_row - 1} rows, {len(headers)} columns")

    field_map = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        hl = h.lower().strip()
        if "full name" in hl or hl == "name":
            field_map["fullName"] = i
        elif "grad" in hl and "year" in hl:
            field_map["gradYear"] = i
        elif "aza" in hl and "bbg" in hl:
            field_map["order"] = i
        elif hl in ("phone number", "phone", "preferred phone"):
            field_map["phone"] = i
        elif hl == "email":
            field_map["email"] = i
        elif "birthdate" in hl or "birth" in hl:
            field_map["birthdate"] = i
        elif "membership status" in hl:
            field_map["membershipStatus"] = i
        elif "join date" in hl or "membership join" in hl:
            field_map["joinDate"] = i
        elif "chapter" in hl and "name" in hl:
            field_map["chapterName"] = i
        elif "parent 1 name" in hl or "parent / legal guardian #1" in hl:
            field_map["parent1Name"] = i
        elif "parent 1 email" in hl:
            field_map["parent1Email"] = i
        elif "parent 1 cell" in hl:
            field_map["parent1Cell"] = i
        elif "parent 2 name" in hl or "parent / legal guardian #2" in hl:
            field_map["parent2Name"] = i
        elif "parent 2 email" in hl:
            field_map["parent2Email"] = i
        elif "parent 2 cell" in hl:
            field_map["parent2Cell"] = i
        elif "recommended" in hl and "program" in hl:
            field_map["recommendedProgram"] = i

    print(f"Mapped fields: {list(field_map.keys())}")

    if "fullName" not in field_map:
        print("ERROR: Could not find a 'Full Name' column. Aborting.")
        sys.exit(1)

    def get(row, field):
        idx = field_map.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def format_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return s if s else None

    def clean_str(val):
        if val is None:
            return None
        s = str(val).strip()
        return s if s and s.lower() != "none" else None

    def clean_phone(val):
        if val is None:
            return None
        s = str(val).strip().replace(".0", "")
        return s if s and s.lower() != "none" else None

    members = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = get(row, "fullName")
        if not name or not str(name).strip():
            continue

        program = clean_str(get(row, "recommendedProgram"))
        if program and program.startswith("="):
            program = program[1:]

        member = {
            "fullName": str(name).strip(),
            "gradYear": clean_str(get(row, "gradYear")),
            "order": clean_str(get(row, "order")),
            "phone": clean_phone(get(row, "phone")),
            "email": clean_str(get(row, "email")),
            "birthdate": format_date(get(row, "birthdate")),
            "membershipStatus": clean_str(get(row, "membershipStatus")),
            "joinDate": format_date(get(row, "joinDate")),
            "chapterName": clean_str(get(row, "chapterName")),
            "parent1Name": clean_str(get(row, "parent1Name")),
            "parent1Email": clean_str(get(row, "parent1Email")),
            "parent1Cell": clean_phone(get(row, "parent1Cell")),
            "parent2Name": clean_str(get(row, "parent2Name")),
            "parent2Email": clean_str(get(row, "parent2Email")),
            "parent2Cell": clean_phone(get(row, "parent2Cell")),
            "recommendedProgram": program,
        }
        members.append(member)

    members.sort(key=lambda m: m["fullName"])
    print(f"Processed {len(members)} members")
    return members


# ---------- Supabase upload ----------

def strip_sensitive(members):
    """Return a copy of members with contact info removed for safe upload."""
    stripped = []
    for m in members:
        clean = {k: v for k, v in m.items() if k not in SENSITIVE_FIELDS}
        stripped.append(clean)
    return stripped


def upload_to_supabase(members):
    """Strip sensitive data and upload JSON to Supabase Storage."""
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        print("\nERROR: Supabase not configured. Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY.")
        sys.exit(1)

    safe_members = strip_sensitive(members)
    json_bytes = json.dumps(safe_members, indent=2).encode("utf-8")

    upload_url = f"{SUPABASE_URL}/storage/v1/object/membership/members.json"

    print(f"\nStripping PII for upload...")
    print(f"Uploading {len(safe_members)} members...")

    req = urllib.request.Request(
        upload_url,
        data=json_bytes,
        method="PUT",
        headers={
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
            "Content-Type": "application/json",
            "x-upsert": "true",
        },
    )

    try:
        with urllib.request.urlopen(req) as resp:
            if resp.status in (200, 201):
                print("Uploaded successfully!")
                return True
            else:
                print(f"Upload returned status {resp.status}")
                return False
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"Upload failed ({e.code}): {body}")
        return False
    except Exception as e:
        print(f"Upload failed: {e}")
        return False


# ---------- Main ----------

def main():
    print("=" * 50)
    print("BBYO Membership Auto-Update")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 50)

    # Validate config
    if not BBYO_EMAIL or not BBYO_PASSWORD:
        print("ERROR: BBYO_EMAIL and BBYO_PASSWORD must be set.")
        sys.exit(1)

    # Step 1: Scrape the export
    xlsx_path = scrape_membership_export()

    # Step 2: Convert to JSON
    members = convert_spreadsheet(xlsx_path)

    # Summary
    chapters = set(m["chapterName"] for m in members if m["chapterName"])
    orders = {}
    for m in members:
        o = m["order"] or "Unknown"
        orders[o] = orders.get(o, 0) + 1

    print(f"\nSummary: {len(members)} members, {len(chapters)} chapters")
    for o, c in sorted(orders.items()):
        print(f"  {o}: {c}")

    # Step 3: Upload to Supabase
    success = upload_to_supabase(members)

    # Cleanup
    try:
        os.remove(xlsx_path)
        print(f"Cleaned up temp file: {xlsx_path}")
    except Exception:
        pass

    if success:
        print("\nDone! The PWA will show updated data on next load.")
    else:
        print("\nUpload failed.")
        sys.exit(1)


if __name__ == "__main__":
    main()
