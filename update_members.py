#!/usr/bin/env python3
"""
Update GCR Membership data from an Excel spreadsheet.

Usage:
    python3 update_members.py                              # uses default spreadsheet path
    python3 update_members.py /path/to/spreadsheet.xlsx    # custom spreadsheet

This will:
  1. Convert the spreadsheet to JSON
  2. Save locally as MemberData.json (bundled fallback)
  3. Upload to Supabase so the app updates automatically (if configured)
"""

import sys
import json
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# ---------- Configuration ----------

DEFAULT_PATH = os.path.expanduser("~/Downloads/Membership.xlsx")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, ".env")

def load_env():
    """Load Supabase config from .env file."""
    config = {}
    if os.path.exists(ENV_PATH):
        with open(ENV_PATH) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    config[key.strip()] = val.strip()
    return config

# ---------- Spreadsheet conversion ----------

def convert_spreadsheet(input_path):
    print(f"Reading: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    print(f"Found {ws.max_row - 1} rows, {len(headers)} columns")

    field_map = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        hl = h.lower().strip()
        if "full name" in hl or hl == "name":
            field_map["fullName"] = i
        elif "grad" in hl and "year" in hl:
            field_map["gradYear"] = i
        elif "aza" in hl and "bbg" in hl:
            field_map["order"] = i
        elif hl in ("phone number", "phone", "preferred phone"):
            field_map["phone"] = i
        elif hl == "email":
            field_map["email"] = i
        elif "birthdate" in hl or "birth" in hl:
            field_map["birthdate"] = i
        elif "membership status" in hl:
            field_map["membershipStatus"] = i
        elif "join date" in hl or "membership join" in hl:
            field_map["joinDate"] = i
        elif "chapter" in hl and "name" in hl:
            field_map["chapterName"] = i
        elif "parent 1 name" in hl or "parent / legal guardian #1" in hl:
            field_map["parent1Name"] = i
        elif "parent 1 email" in hl:
            field_map["parent1Email"] = i
        elif "parent 1 cell" in hl:
            field_map["parent1Cell"] = i
        elif "parent 2 name" in hl or "parent / legal guardian #2" in hl:
            field_map["parent2Name"] = i
        elif "parent 2 email" in hl:
            field_map["parent2Email"] = i
        elif "parent 2 cell" in hl:
            field_map["parent2Cell"] = i
        elif "recommended" in hl and "program" in hl:
            field_map["recommendedProgram"] = i

    print(f"Mapped fields: {list(field_map.keys())}")

    if "fullName" not in field_map:
        print("ERROR: Could not find a 'Full Name' column. Aborting.")
        sys.exit(1)

    def get(row, field):
        idx = field_map.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def format_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return s if s else None

    def clean_str(val):
        if val is None:
            return None
        s = str(val).strip()
        return s if s and s.lower() != "none" else None

    def clean_phone(val):
        if val is None:
            return None
        s = str(val).strip().replace(".0", "")
        return s if s and s.lower() != "none" else None

    members = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = get(row, "fullName")
        if not name or not str(name).strip():
            continue

        program = clean_str(get(row, "recommendedProgram"))
        if program and program.startswith("="):
            program = program[1:]

        member = {
            "fullName": str(name).strip(),
            "gradYear": clean_str(get(row, "gradYear")),
            "order": clean_str(get(row, "order")),
            "phone": clean_phone(get(row, "phone")),
            "email": clean_str(get(row, "email")),
            "birthdate": format_date(get(row, "birthdate")),
            "membershipStatus": clean_str(get(row, "membershipStatus")),
            "joinDate": format_date(get(row, "joinDate")),
            "chapterName": clean_str(get(row, "chapterName")),
            "parent1Name": clean_str(get(row, "parent1Name")),
            "parent1Email": clean_str(get(row, "parent1Email")),
            "parent1Cell": clean_phone(get(row, "parent1Cell")),
            "parent2Name": clean_str(get(row, "parent2Name")),
            "parent2Email": clean_str(get(row, "parent2Email")),
            "parent2Cell": clean_phone(get(row, "parent2Cell")),
            "recommendedProgram": program,
        }
        members.append(member)

    members.sort(key=lambda m: m["fullName"])
    return members

# ---------- Supabase upload ----------

SENSITIVE_FIELDS = [
    "phone", "email",
    "parent1Name", "parent1Email", "parent1Cell",
    "parent2Name", "parent2Email", "parent2Cell",
]

def strip_sensitive(members):
    """Return a copy of members with contact info removed for safe upload."""
    stripped = []
    for m in members:
        clean = {k: v for k, v in m.items() if k not in SENSITIVE_FIELDS}
        stripped.append(clean)
    return stripped

def upload_to_supabase(members, config):
    """Strip sensitive data and upload JSON to Supabase Storage via REST API."""
    import urllib.request

    supabase_url = config.get("SUPABASE_URL", "")
    service_key = config.get("SUPABASE_SERVICE_ROLE_KEY", "")

    if not supabase_url or not service_key:
        print("\nSupabase not configured. Skipping upload.")
        print("To enable, create a .env file with:")
        print("  SUPABASE_URL=https://your-project.supabase.co")
        print("  SUPABASE_SERVICE_ROLE_KEY=your-service-role-key")
        return False

    safe_members = strip_sensitive(members)
    json_bytes = json.dumps(safe_members, indent=2).encode("utf-8")

    upload_url = f"{supabase_url}/storage/v1/object/membership/members.json"

    print(f"\nStripping emails & phone numbers for upload...")
    print(f"Uploading {len(safe_members)} members (contact info removed)...")

    req = urllib.request.Request(
        upload_url,
        data=json_bytes,
        method="PUT",
        headers={
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
            "x-upsert": "true",
        },
    )

    try:
        with urllib.request.urlopen(req) as resp:
            if resp.status in (200, 201):
                print("Uploaded successfully!")
                print("(Bucket is private — data is not publicly accessible)")
                return True
            else:
                print(f"Upload returned status {resp.status}")
                return False
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"Upload failed ({e.code}): {body}")
        return False
    except Exception as e:
        print(f"Upload failed: {e}")
        return False

# ---------- Main ----------

def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    if not os.path.exists(path):
        print(f"File not found: {path}")
        print(f"Usage: python3 {sys.argv[0]} /path/to/spreadsheet.xlsx")
        sys.exit(1)

    # Convert spreadsheet
    members = convert_spreadsheet(path)
    json_str = json.dumps(members, indent=2)
    json_bytes = json_str.encode("utf-8")

    # Save locally (bundled fallback)
    output_path = os.path.join(SCRIPT_DIR, "GCR Membership", "MemberData.json")
    with open(output_path, "w") as f:
        f.write(json_str)

    # Summary
    chapters = set(m["chapterName"] for m in members if m["chapterName"])
    orders = {}
    for m in members:
        o = m["order"] or "Unknown"
        orders[o] = orders.get(o, 0) + 1

    print(f"\nExported {len(members)} members")
    print(f"Chapters: {len(chapters)}")
    for o, c in sorted(orders.items()):
        print(f"  {o}: {c}")

    # Upload to Supabase (with sensitive data stripped)
    config = load_env()
    uploaded = upload_to_supabase(members, config)

    if uploaded:
        print("\nDone! The app will show updated data on next launch.")
    else:
        print("\nLocal JSON updated. Rebuild in Xcode (Cmd+R) to update bundled data.")

if __name__ == "__main__":
    main()
